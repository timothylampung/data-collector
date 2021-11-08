if __name__ == '__main__':
    from openpyxl import Workbook, load_workbook
    import serial
    import time
    from threading import Timer

    arduino = serial.Serial(port='/dev/cu.usbmodem11301', baudrate=9600, timeout=.1)

    sentinel = None
    stop_timer = False
    duration = 300
    worksheet_name = 'WITH PROBE EMPTY'
    filename = 'no_spin_water_fan_temperature_capture.xlsx'

    try:
        wb = load_workbook(filename)
    except:
        wb = Workbook()

    exist = False
    for _ws in wb.worksheets:
        if _ws.title == worksheet_name:
            raise Exception('worksheet exist')

    ws = wb.create_sheet(worksheet_name)
    TIME = 'A'
    TEMPERATURE = 'B'
    TEMPERATURE_IDLE = 'C'
    TEMPERATURE_A = 'D'
    TEMPERATURE_T = 'E'

    ws[f'{TEMPERATURE_T}1'] = 'Temperature Thermistor (celsius)'
    ws[f'{TEMPERATURE}1'] = 'Temperature spin (celsius)'
    ws[f'{TEMPERATURE_IDLE}1'] = 'Temperature Idle (celsius)'
    ws[f'{TEMPERATURE_A}1'] = 'Ambient Temperature (celsius)'
    ws[f'{TIME}1'] = 'Time (seconds)'

    wb.save(filename)


    def write_read():
        data = arduino.readline()
        data = data.decode('utf-8')
        return data


    index = 1
    state = 'IDLE'

    while not stop_timer:
        value = write_read()
        if value is not None and value != '':
            value = value.strip()
            if sentinel == 'BOOTED UP!':
                index = index + 1
                # IDLE | 61 | 33.05 | -1 | 32.37
                state, time, temp, probe_t, temp_a = tuple(value.split('|'))
                print(state, temp, time, temp_a, probe_t)

                state = state
                time = float(time)
                temp = float(temp)
                temp_a = float(temp_a)
                probe_t = float(probe_t)
                if state == 'IDLE':
                    ws[f'{TEMPERATURE_IDLE}{index}'] = f'{temp}'
                else:
                    ws[f'{TEMPERATURE}{index}'] = f'{temp}'

                ws[f'{TEMPERATURE_A}{index}'] = f'{temp_a}'
                ws[f'{TIME}{index}'] = f'{time}'
                ws[f'{TEMPERATURE_T}{index}'] = f'{probe_t}'

                if time > duration:
                    stop_timer = True

            else:
                pass
                sentinel = value

    print('DONE!')
    wb.save(filename)
